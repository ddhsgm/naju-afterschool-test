from __future__ import annotations

import json
import re
from datetime import datetime
from pathlib import Path

import openpyxl


BASE_DIR = Path(__file__).resolve().parent
ROOT_DIR = BASE_DIR.parent
SOURCE_DIR = ROOT_DIR / "원본자료"
STUDENT_XLSX = SOURCE_DIR / "학생명단(3.16.기준).xlsx"
ROOM_XLSX = SOURCE_DIR / "2026. 초등돌봄·방과후 운영교실.xlsx"
OUTPUT_JS = BASE_DIR / "data.js"
BOOTSTRAP_JSON = BASE_DIR / "bootstrap.json"
MISSING_CONTACTS = BASE_DIR / "missing-contacts.txt"
MISSING_CONTACTS_JSON = BASE_DIR / "missing-contacts.json"
SUPABASE_SQL = BASE_DIR / "supabase-seed.sql"
CARE_PATTERN = "*돌봄교실 학생명단*.xlsx"
FREE_PASS_PATTERN = "*자유수강권*.xlsx"

PERIODS = {
    "1부": {"start": "13:25", "end": "14:05"},
    "2부": {"start": "14:10", "end": "14:50"},
    "3부": {"start": "14:55", "end": "15:35"},
    "4부": {"start": "15:40", "end": "16:20"},
}

COURSE_DEFINITIONS = [
    {
        "id": "creative-art",
        "name": "창의미술",
        "days": ["화", "금"],
        "room_key": "창의미술",
        "teacher": "임형은",
        "teacherPhone": "010-7765-2870",
        "capacity": 20,
        "feeType": "수익자",
        "fee": "35,000원",
        "note": "매월 10,000원",
        "slots": [
            {"period": "2부", "gradeMin": 1, "gradeMax": 4},
            {"period": "3부", "gradeMin": 1, "gradeMax": 6},
            {"period": "4부", "gradeMin": 1, "gradeMax": 6},
        ],
    },
    {
        "id": "creative-robot",
        "name": "창의로봇",
        "days": ["화", "금"],
        "room_key": "창의로봇",
        "teacher": "서원희",
        "teacherPhone": "010-2661-0362",
        "capacity": 20,
        "feeType": "수익자",
        "fee": "35,000원",
        "note": "재료비 수준별 88,000~99,000원",
        "slots": [
            {"period": "2부", "gradeMin": 1, "gradeMax": 4},
            {"period": "3부", "gradeMin": 1, "gradeMax": 6},
            {"period": "4부", "gradeMin": 1, "gradeMax": 6},
        ],
    },
    {
        "id": "computer",
        "name": "컴퓨터",
        "days": ["월", "목"],
        "room_key": "컴퓨터",
        "teacher": "홍설희",
        "teacherPhone": "010-2890-8683",
        "capacity": 25,
        "feeType": "수익자",
        "fee": "35,000원",
        "note": "교재비 수준별 12,000~18,000원",
        "slots": [
            {"period": "1부", "gradeMin": 1, "gradeMax": 2},
            {"period": "2부", "gradeMin": 3, "gradeMax": 4},
            {"period": "3부", "gradeMin": 1, "gradeMax": 6},
        ],
    },
    {
        "id": "abacus",
        "name": "주산",
        "days": ["월", "목"],
        "room_key": "주산",
        "teacher": "김경매",
        "teacherPhone": "010-6406-5336",
        "capacity": 20,
        "feeType": "수익자",
        "fee": "35,000원",
        "note": "주판 13,000원, 교재비 수준별 8,000~11,000원",
        "slots": [
            {"period": "1부", "gradeMin": 1, "gradeMax": 2},
            {"period": "2부", "gradeMin": 1, "gradeMax": 4},
            {"period": "3부", "gradeMin": 1, "gradeMax": 6},
        ],
    },
    {
        "id": "hanja",
        "name": "한자",
        "days": ["월", "금"],
        "room_key": "한자",
        "teacher": "심희숙",
        "teacherPhone": "010-5581-6733",
        "capacity": 20,
        "feeType": "수익자",
        "fee": "35,000원",
        "note": "교재비 수준별 12,000~16,000원",
        "slots": [
            {"period": "2부", "gradeMin": 1, "gradeMax": 4},
            {"period": "3부", "gradeMin": 1, "gradeMax": 6},
            {"period": "4부", "gradeMin": 1, "gradeMax": 6},
        ],
    },
    {
        "id": "piano",
        "name": "피아노",
        "days": ["화", "금"],
        "room_key": "피아노",
        "teacher": "강민정",
        "teacherPhone": "010-6800-4233",
        "capacity": 14,
        "feeType": "무료",
        "fee": "무료",
        "note": "교재비 수준별 5,000~8,000원",
        "slots": [
            {"period": "2부", "gradeMin": 1, "gradeMax": 4},
            {"period": "3부", "gradeMin": 1, "gradeMax": 6},
            {"period": "4부", "gradeMin": 1, "gradeMax": 6},
        ],
    },
    {
        "id": "violin",
        "name": "바이올린",
        "days": ["화", "금"],
        "room_key": "바이올린",
        "teacher": "김규리",
        "teacherPhone": "010-4180-9863",
        "capacity": 12,
        "feeType": "무료",
        "fee": "무료",
        "note": "교재비 수준별 4,000~12,000원",
        "slots": [
            {"period": "2부", "gradeMin": 1, "gradeMax": 4},
            {"period": "3부", "gradeMin": 1, "gradeMax": 6},
        ],
    },
    {
        "id": "jump-rope",
        "name": "줄넘기",
        "days": ["월", "목"],
        "room_key": "줄넘기",
        "teacher": "강선영",
        "teacherPhone": "010-8512-4770",
        "capacity": 20,
        "feeType": "무료",
        "fee": "무료",
        "note": "개인 줄넘기, 운동화, 물병 준비",
        "slots": [
            {"period": "1부", "gradeMin": 1, "gradeMax": 2},
            {"period": "2부", "gradeMin": 1, "gradeMax": 4},
            {"period": "3부", "gradeMin": 1, "gradeMax": 6},
        ],
    },
    {
        "id": "table-tennis",
        "name": "탁구",
        "days": ["화", "금"],
        "room_key": "탁구",
        "teacher": "이지우",
        "teacherPhone": "010-8603-0030",
        "capacity": 12,
        "feeType": "무료",
        "fee": "무료",
        "note": "맞춤형 프로그램",
        "slots": [
            {"period": "2부", "gradeMin": 1, "gradeMax": 4},
            {"period": "3부", "gradeMin": 1, "gradeMax": 6},
        ],
    },
    {
        "id": "dance",
        "name": "방송댄스",
        "days": ["월", "목"],
        "room_key": "방송댄스",
        "teacher": "주연지",
        "teacherPhone": "010-8200-2978",
        "capacity": 20,
        "feeType": "무료",
        "fee": "무료",
        "note": "2부 무용실 운영",
        "slots": [
            {"period": "2부", "gradeMin": 1, "gradeMax": 2},
            {"period": "3부", "gradeMin": 3, "gradeMax": 6},
        ],
    },
    {
        "id": "new-sports",
        "name": "뉴스포츠",
        "days": ["금"],
        "room_key": "뉴스포츠",
        "teacher": "박종현",
        "teacherPhone": "010-5047-3281",
        "capacity": 20,
        "feeType": "무료",
        "fee": "무료",
        "note": "금요일 단일 운영",
        "slots": [
            {"period": "2부", "gradeMin": 3, "gradeMax": 4},
            {"period": "3부", "gradeMin": 5, "gradeMax": 6},
        ],
    },
    {
        "id": "book-play",
        "name": "책놀이",
        "days": ["월", "목"],
        "room_key": "책놀이",
        "teacher": "이향현",
        "teacherPhone": "010-8604-3197",
        "capacity": 15,
        "feeType": "무료",
        "fee": "무료",
        "note": "도서관 운영",
        "slots": [
            {"period": "1부", "gradeMin": 1, "gradeMax": 2},
            {"period": "2부", "gradeMin": 3, "gradeMax": 4},
            {"period": "3부", "gradeMin": 5, "gradeMax": 6},
        ],
    },
    {
        "id": "math-play",
        "name": "놀이수학",
        "days": ["월", "목"],
        "room_key": "놀이수학",
        "teacher": "최나라",
        "teacherPhone": "010-3163-7227",
        "capacity": 16,
        "feeType": "무료",
        "fee": "무료",
        "note": "맞춤형 프로그램",
        "slots": [
            {"period": "2부", "gradeMin": 1, "gradeMax": 2},
        ],
    },
    {
        "id": "jamjam-english",
        "name": "잼잼잉글리쉬",
        "days": ["화", "금"],
        "room_key": "잼잼잉글리쉬(2부)",
        "teacher": "김효빈 / 이지영",
        "teacherPhone": "010-8001-3655 / 010-4823-6104",
        "capacity": 16,
        "feeType": "무료",
        "fee": "무료",
        "note": "2부 교담수업실 1, 3부 북카페 운영",
        "slots": [
            {"period": "2부", "gradeMin": 1, "gradeMax": 2, "roomKey": "잼잼잉글리쉬(2부)"},
            {"period": "3부", "gradeMin": 1, "gradeMax": 2, "roomKey": "잼잼잉글리쉬(3부)"},
        ],
    },
    {
        "id": "play-physical",
        "name": "놀이체육",
        "days": ["수"],
        "room_key": "놀이체육",
        "teacher": "박종현",
        "teacherPhone": "010-5047-3281",
        "capacity": 18,
        "feeType": "무료",
        "fee": "무료",
        "note": "수요일 단일 운영",
        "slots": [
            {"period": "2부", "gradeMin": 1, "gradeMax": 2},
            {"period": "3부", "gradeMin": 1, "gradeMax": 2},
        ],
    },
]


def normalize_phone(value: object) -> str:
    return re.sub(r"\D", "", str(value or ""))


def extract_number(value: object) -> int:
    matched = re.search(r"(\d+)", str(value or ""))
    return int(matched.group(1)) if matched else 0


def sql_literal(value: object) -> str:
    if value is None:
        return "NULL"
    return "'" + str(value).replace("'", "''") + "'"


def find_optional_source(pattern: str) -> Path | None:
    direct_matches = sorted(SOURCE_DIR.glob(pattern))
    if direct_matches:
        return direct_matches[0]

    parent_dir = ROOT_DIR.parent
    fallback_matches = sorted(parent_dir.rglob(pattern))
    if fallback_matches:
        return fallback_matches[0]

    return None


def build_room_index() -> dict[str, dict[str, str]]:
    workbook = openpyxl.load_workbook(ROOM_XLSX, data_only=True)
    sheet = workbook.active
    room_index: dict[str, dict[str, str]] = {}
    current_room_name = ""
    current_location = ""

    for row in sheet.iter_rows(min_row=4, values_only=True):
        _, room_name, location, _, program_name, days, _, note = row
        if room_name:
            current_room_name = str(room_name).strip()
        if location:
            current_location = str(location).strip()
        if not program_name:
            continue
        room_index[str(program_name).strip()] = {
            "room": current_room_name or "",
            "location": current_location or "",
            "days": str(days or "").strip(),
            "note": str(note or "").strip(),
        }

    return room_index


def build_students() -> tuple[list[dict[str, object]], list[str]]:
    workbook = openpyxl.load_workbook(STUDENT_XLSX, data_only=True)
    sheet = workbook["E알리미"]

    students: list[dict[str, object]] = []
    missing_contacts: list[str] = []

    for row in sheet.iter_rows(min_row=3, values_only=True):
        name = str(row[0] or "").strip()
        if not name:
            continue

        grade = str(row[1] or "").strip()
        class_room = str(row[2] or "").strip()
        number = str(row[3] or "").strip()
        primary_phone = normalize_phone(row[4])
        secondary_phone = normalize_phone(row[5])
        reply_phone = normalize_phone(row[7])
        phone = primary_phone or secondary_phone or reply_phone

        grade_number = extract_number(grade)
        class_number = extract_number(class_room)
        if not phone:
            missing_contacts.append(f"{grade} {class_room} {name}")
            continue

        student_id = f"{grade_number}-{class_number}-{name}"
        students.append(
            {
                "id": student_id,
                "name": name,
                "gradeLabel": grade,
                "classLabel": class_room,
                "numberLabel": number,
                "grade": grade_number,
                "classRoom": class_number,
                "phone": phone,
            }
        )

    students.sort(key=lambda item: (item["grade"], item["classRoom"], item["name"]))
    missing_contacts.sort()
    return students, missing_contacts


def build_courses(room_index: dict[str, dict[str, str]]) -> list[dict[str, object]]:
    courses: list[dict[str, object]] = []

    for course in COURSE_DEFINITIONS:
        room_info = room_index.get(course["room_key"], {})
        slots = []
        for slot in course["slots"]:
            room_key = slot.get("roomKey", course["room_key"])
            slot_room_info = room_index.get(str(room_key), room_info)
            period = str(slot["period"])
            slots.append(
                {
                    "id": f"{course['id']}::{period}",
                    "period": period,
                    "start": PERIODS[period]["start"],
                    "end": PERIODS[period]["end"],
                    "gradeMin": slot["gradeMin"],
                    "gradeMax": slot["gradeMax"],
                    "days": course["days"],
                    "room": slot_room_info.get("room", ""),
                    "location": slot_room_info.get("location", ""),
                    "capacity": course["capacity"],
                }
            )

        courses.append(
            {
                "id": course["id"],
                "name": course["name"],
                "days": course["days"],
                "room": room_info.get("room", ""),
                "location": room_info.get("location", ""),
                "teacher": course["teacher"],
                "teacherPhone": course["teacherPhone"],
                "capacity": course["capacity"],
                "feeType": course["feeType"],
                "fee": course["fee"],
                "note": course["note"],
                "slots": slots,
            }
        )

    return courses


def build_care_map() -> dict[str, str]:
    care_path = find_optional_source(CARE_PATTERN)
    if not care_path:
        return {}

    workbook = openpyxl.load_workbook(care_path, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    care_map: dict[str, str] = {}

    for row in sheet.iter_rows(min_row=4, values_only=True):
        grade, class_room, name = row[3], row[4], row[6]
        care_class = str(row[2] or "").strip()
        if not grade or not class_room or not name or not care_class:
            continue

        student_id = f"{extract_number(grade)}-{extract_number(class_room)}-{str(name).strip()}"
        care_map[student_id] = care_class

    return care_map


def build_free_pass_map(students: list[dict[str, object]]) -> dict[str, dict[str, object]]:
    free_pass_path = find_optional_source(FREE_PASS_PATTERN)
    if not free_pass_path:
        return {}

    workbook = openpyxl.load_workbook(free_pass_path, data_only=True)
    support_sheet = None
    for sheet in workbook.worksheets:
        if "지원명단" in sheet.title:
            support_sheet = sheet
            break

    if support_sheet is None:
        return {}

    students_by_name = {str(student["name"]).strip(): student for student in students}
    free_pass_map: dict[str, dict[str, object]] = {}

    for row in support_sheet.iter_rows(min_row=4, values_only=True):
        _, old_grade, _, _, name, *monthly_amounts = row
        student_name = str(name or "").strip()
        if not student_name:
            continue

        current_student = students_by_name.get(student_name)
        if not current_student:
            continue

        old_grade_number = extract_number(old_grade)
        current_grade_number = int(current_student["grade"])

        if old_grade_number <= 0 or old_grade_number >= 6:
            continue
        if current_grade_number == 1:
            continue
        if current_grade_number != old_grade_number + 1:
            continue

        support_total = sum(int(value or 0) for value in monthly_amounts if isinstance(value, (int, float)))
        free_pass_map[str(current_student["id"])] = {
            "matchedFromYear": 2025,
            "supportTotal": support_total,
            "sourceGrade": old_grade_number,
            "currentGrade": current_grade_number,
        }

    return free_pass_map


def build_supabase_sql(students: list[dict[str, object]], courses: list[dict[str, object]]) -> str:
    lines = [
        "-- Generated by build_data.py",
        "begin;",
        "",
        "create table if not exists public.students (",
        "  id text primary key,",
        "  name text not null,",
        "  grade integer not null,",
        "  grade_label text not null,",
        "  class_room integer not null,",
        "  class_label text not null,",
        "  number_label text,",
        "  phone text not null,",
        "  active boolean not null default true",
        ");",
        "",
        "create table if not exists public.courses (",
        "  id text primary key,",
        "  name text not null,",
        "  days_json jsonb not null,",
        "  room text,",
        "  location text,",
        "  teacher text,",
        "  teacher_phone text,",
        "  capacity integer,",
        "  fee_type text,",
        "  fee text,",
        "  note text",
        ");",
        "",
        "create table if not exists public.course_slots (",
        "  id text primary key,",
        "  course_id text not null references public.courses(id) on delete cascade,",
        "  period text not null,",
        "  start_time text not null,",
        "  end_time text not null,",
        "  grade_min integer not null,",
        "  grade_max integer not null,",
        "  days_json jsonb not null,",
        "  room text,",
        "  location text,",
        "  capacity integer",
        ");",
        "",
        "create table if not exists public.applications (",
        "  student_id text not null references public.students(id) on delete cascade,",
        "  slot_id text not null references public.course_slots(id) on delete cascade,",
        "  updated_at timestamptz not null default now(),",
        "  primary key (student_id, slot_id)",
        ");",
        "",
        "alter table public.students enable row level security;",
        "alter table public.courses enable row level security;",
        "alter table public.course_slots enable row level security;",
        "alter table public.applications enable row level security;",
        "",
        "drop policy if exists deny_all_students on public.students;",
        "drop policy if exists deny_all_courses on public.courses;",
        "drop policy if exists deny_all_course_slots on public.course_slots;",
        "drop policy if exists deny_all_applications on public.applications;",
        "",
        "create policy deny_all_students on public.students for all using (false) with check (false);",
        "create policy deny_all_courses on public.courses for all using (false) with check (false);",
        "create policy deny_all_course_slots on public.course_slots for all using (false) with check (false);",
        "create policy deny_all_applications on public.applications for all using (false) with check (false);",
        "",
        "truncate table public.applications;",
        "truncate table public.course_slots cascade;",
        "truncate table public.courses cascade;",
        "truncate table public.students cascade;",
        "",
    ]

    for student in students:
        lines.append(
            "insert into public.students (id, name, grade, grade_label, class_room, class_label, number_label, phone, active) values "
            f"({sql_literal(student['id'])}, {sql_literal(student['name'])}, {student['grade']}, {sql_literal(student['gradeLabel'])}, "
            f"{student['classRoom']}, {sql_literal(student['classLabel'])}, {sql_literal(student['numberLabel'])}, {sql_literal(student['phone'])}, true);"
        )

    lines.append("")

    for course in courses:
        lines.append(
            "insert into public.courses (id, name, days_json, room, location, teacher, teacher_phone, capacity, fee_type, fee, note) values "
            f"({sql_literal(course['id'])}, {sql_literal(course['name'])}, {sql_literal(json.dumps(course['days'], ensure_ascii=False))}::jsonb, "
            f"{sql_literal(course['room'])}, {sql_literal(course['location'])}, {sql_literal(course['teacher'])}, "
            f"{sql_literal(course['teacherPhone'])}, {course['capacity']}, {sql_literal(course['feeType'])}, {sql_literal(course['fee'])}, {sql_literal(course['note'])});"
        )

    lines.append("")

    for course in courses:
        for slot in course["slots"]:
            lines.append(
                "insert into public.course_slots (id, course_id, period, start_time, end_time, grade_min, grade_max, days_json, room, location, capacity) values "
                f"({sql_literal(slot['id'])}, {sql_literal(course['id'])}, {sql_literal(slot['period'])}, {sql_literal(slot['start'])}, {sql_literal(slot['end'])}, "
                f"{slot['gradeMin']}, {slot['gradeMax']}, {sql_literal(json.dumps(slot['days'], ensure_ascii=False))}::jsonb, "
                f"{sql_literal(slot['room'])}, {sql_literal(slot['location'])}, {slot['capacity']});"
            )

    lines.extend(["", "commit;", ""])
    return "\n".join(lines)


def write_outputs() -> None:
    room_index = build_room_index()
    students, missing_contacts = build_students()
    courses = build_courses(room_index)
    care_map = build_care_map()
    free_pass_map = build_free_pass_map(students)

    meta = {
        "schoolName": "나주중앙초등학교",
        "generatedAt": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "studentCount": len(students),
        "missingContactCount": len(missing_contacts),
        "periods": PERIODS,
        "days": ["월", "화", "수", "목", "금"],
        "storageKey": "naju-jungang-afterschool-applications-v1",
        "maxGrade": max((student["grade"] for student in students), default=6),
        "maxClass": max((student["classRoom"] for student in students), default=6),
    }
    payload = {"meta": meta, "students": students, "courses": courses}
    bootstrap_payload = {
        "meta": meta,
        "courses": courses,
        "support": {
            "careByStudentId": care_map,
            "freePassByStudentId": free_pass_map,
            "voucherGrades": [3],
        },
    }

    OUTPUT_JS.write_text(
        "window.AFTERSCHOOL_DATA = " + json.dumps(payload, ensure_ascii=False, indent=2) + ";\n",
        encoding="utf-8",
    )
    BOOTSTRAP_JSON.write_text(json.dumps(bootstrap_payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    MISSING_CONTACTS.write_text(
        "\n".join(
            [
                "연락처가 없어 로그인 대상에서 제외된 학생 목록",
                f"총 {len(missing_contacts)}명",
                "",
                *missing_contacts,
            ]
        )
        + "\n",
        encoding="utf-8",
    )
    MISSING_CONTACTS_JSON.write_text(json.dumps(missing_contacts, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    SUPABASE_SQL.write_text(build_supabase_sql(students, courses), encoding="utf-8")

    print(
        json.dumps(
            {
                "students": len(students),
                "missing_contacts": len(missing_contacts),
                "courses": len(courses),
                "output": str(OUTPUT_JS),
                "bootstrap": str(BOOTSTRAP_JSON),
                "sql": str(SUPABASE_SQL),
            },
            ensure_ascii=False,
        )
    )


if __name__ == "__main__":
    write_outputs()
